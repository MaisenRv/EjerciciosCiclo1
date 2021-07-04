import openpyxl

#direccion del archivo
archivo = openpyxl.load_workbook("archivo.xlsx")

#selecciona la hoja del excel
h1 = archivo["Hoja1"]


cedula = int(input("cedula: "))

#seleccion de una celda en especifico
#row es la fila
#column es la columna
dato = h1.cell(row=10, column=3).value

print(dato)
existe = False
#max_row es la ultuma fila que este llena
#max_column es la ultima columna llena

for fila in range(2,h1.max_row):
    ced = h1.cell(row=fila,column=1).value

    if ced == cedula:
        apellido = h1.cell(row=fila,column=2).value
        nombre = h1.cell(row=fila,column=3).value
        fecha = h1.cell(row=fila,column=4).value[:10]

        print("Nombre: ",nombre)
        print("Fecha:",fecha)
        break

if not existe:
    print("cedula no registrada")