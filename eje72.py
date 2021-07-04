from typing import AsyncIterable
import openpyxl

#direccion del archivo
archivo = openpyxl.load_workbook("archivo.xlsx")

#selecciona la hoja del excel
h1 = archivo["Hoja1"]

area = int(input("cedula: "))

contador = 0
#max_row es la ultuma fila que este llena
#max_column es la ultima columna llena

#sumar uno para que todas las filas 
#las pueda contar bien

for fila in range(2,h1.max_row + 1):
    area2 = h1.cell(row=fila,column=6).value

    if area.lower() == area2.lower():
        contador += 0
       

if contador == 0:
    print("no hay registros en esta area")

#lo pone en formato de miles
print("{0:,}".format(contador))