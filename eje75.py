import  openpyxl

archivo = openpyxl.load_workbook("archivo.xlsx", data_only=True)
#el data_only solo tiene en cuenta los datos 
#no toma encuenta los formulas 

H1 = archivo["Hoja1"]

#crea una hoja nueva
archivo.create_sheet("Contabilidad")
HC = archivo["Contabilidad"]

area = input("Area: ")

for fila in range(2,H1.max_row):
    area2 =H1.cell(row = fila , column =6).value 
    if area.lower() == area2.lower():
         cedula = H1.cell(row = fila , column =1).value



archivo.save("archivo.xlsx")