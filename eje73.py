from typing import AsyncIterable
import openpyxl

#direccion del archivo
archivo = openpyxl.load_workbook("archivo.xlsx")

#selecciona la hoja del excel
h1 = archivo["Hoja1"]


existe = False
#max_row es la ultuma fila que este llena
#max_column es la ultima columna llena

#sumar uno para que todas las filas 
#las pueda contar bien

for fila in range(2,h1.max_row + 1):
    fecha = str(h1.cell(row=fila,column=4).value)[:4]

    if  fecha < "1980":
        existe = True
        apellido = h1.cell(row=fila,column=2).value
        nombre = h1.cell(row=fila,column=3).value
        cedula = h1.cell(row=fila,column=1).value

        linea = "{0:,}".format(cedula) +"\t" + apellido+ " "+ nombre      
        print(linea)

if not existe:
    print("no hay registros en esta area")

